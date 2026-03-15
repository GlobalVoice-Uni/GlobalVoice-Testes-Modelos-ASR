"""
importar_resultados.py
Importa resultados de benchmarks (JSONs) para a planilha resultados_gerais.xlsx.

Uso:
  python importar_resultados.py --folder resultados/faster-whisper/cpu  
  python importar_resultados.py --folder resultados/         # busca recursiva
  python importar_resultados.py --folder resultados/faster-whisper/cpu --dry-run    # simula sem gravar

Colunas gravadas:
  F  Tempo_Processamento_s
    G  RTF
  H  WER
  L  Pico_RAM_MB
  M  Pico_VRAM_MB
"""

import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Configurações fixas
# ---------------------------------------------------------------------------

WORKBOOK_PATH = Path(__file__).parent / "resultados_gerais.xlsx"
SHEET_NAME = "in"

# Colunas (1-indexed) que serao gravadas
COL_TEMPO = 6   # Tempo_Processamento_s
COL_RTF = 7     # RTF
COL_WER = 8     # WER
# COL 9-11 = Chunks → nao importados
COL_RAM = 12    # Pico_RAM_MB
COL_VRAM = 13   # Pico_VRAM_MB

# Mapeamentos JSON → planilha
MODEL_MAP = {
    "faster-whisper": "Faster-Whisper",
    "openai-whisper": "OpenAI-Whisper",
    "whisperx":       "WhisperX",
}

DEVICE_MAP = {
    "cpu":  "CPU",
    "cuda": "GPU",
}

LANG_MAP = {
    "en": "EN",
    "pt": "BR",
}

# Intervalos canônicos para snap (valor real -> valor da planilha)
CANONICAL_INTERVALS = [10, 30, 60]
INTERVAL_TOLERANCE = 10  # segundos de tolerância para o snap


# ---------------------------------------------------------------------------
# Funções auxiliares
# ---------------------------------------------------------------------------

def snap_interval(audio_duration_s: float) -> int | None:
    """Arredonda a duraçao real do áudio para o intervalo canônico mais próximo."""
    best = min(CANONICAL_INTERVALS, key=lambda c: abs(c - audio_duration_s))
    if abs(best - audio_duration_s) <= INTERVAL_TOLERANCE:
        return best
    return None


def build_key_index(ws) -> dict:
    """
    Lê a planilha e devolve um dict:
        (modelo, hardware, tamanho, idioma, intervalo) -> row_number
    """
    index = {}
    for row in range(2, ws.max_row + 1):
        modelo    = ws.cell(row, 1).value
        hardware  = ws.cell(row, 2).value
        tamanho   = ws.cell(row, 3).value
        idioma    = ws.cell(row, 4).value
        intervalo = ws.cell(row, 5).value

        if not all([modelo, hardware, tamanho, idioma, intervalo]):
            continue

        key = (
            str(modelo).strip(),
            str(hardware).strip(),
            str(tamanho).strip(),
            str(idioma).strip(),
            int(intervalo),
        )
        index[key] = row

    return index


def json_to_key(data: dict) -> tuple | None:
    """
    Converte os campos do JSON na chave (5-tupla) usada no índice da planilha.
    Devolve None se algum campo obrigatório estiver ausente ou nao mapeável.
    """
    model_raw  = data.get("model", "")
    device_raw = data.get("device", "")
    size_raw   = data.get("model_size", "")
    lang_raw   = data.get("language", "")
    dur_raw    = data.get("audio_duration_s")

    modelo    = MODEL_MAP.get(model_raw.lower())
    hardware  = DEVICE_MAP.get(device_raw.lower())
    tamanho   = size_raw.lower() if size_raw else None
    idioma    = LANG_MAP.get(lang_raw.lower())
    intervalo = snap_interval(float(dur_raw)) if dur_raw is not None else None

    if not all([modelo, hardware, tamanho, idioma, intervalo]):
        return None

    return (modelo, hardware, tamanho, idioma, intervalo)


def safe_round(value, ndigits: int = 4):
    """Arredonda apenas se for número."""
    try:
        return round(float(value), ndigits)
    except (TypeError, ValueError):
        return None


def extract_rtf(data: dict) -> float | None:
    """Lê RTF do JSON; se nao houver, calcula por tempo/duraçao quando possível."""
    for key in ("real_time_factor", "rtf"):
        rtf_val = safe_round(data.get(key), 4)
        if rtf_val is not None:
            return rtf_val

    processing_time = safe_round(data.get("processing_time_s"), 8)
    duration = safe_round(data.get("audio_duration_s"), 8)
    if processing_time is None or duration in (None, 0):
        return None

    return safe_round(processing_time / duration, 4)


def backup_workbook(workbook_path: Path) -> Path:
    """Cria backup com timestamp antes de qualquer escrita."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = workbook_path.with_name(f"{workbook_path.stem}_backup_{ts}{workbook_path.suffix}")
    shutil.copy2(workbook_path, backup_path)
    return backup_path


# ---------------------------------------------------------------------------
# Núcleo da importaçao
# ---------------------------------------------------------------------------

def import_folder(folder: Path, dry_run: bool = False):
    json_files = sorted(folder.rglob("*.json"))

    if not json_files:
        print(f"Nenhum JSON encontrado em: {folder}")
        return

    print(f"\n{'[DRY-RUN] ' if dry_run else ''}Importando {len(json_files)} arquivo(s) de: {folder}")
    print(f"Planilha: {WORKBOOK_PATH}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERRO: Aba '{SHEET_NAME}' nao encontrada na planilha.")
        return

    ws = wb[SHEET_NAME]
    key_index = build_key_index(ws)

    stats = {"inseridos": 0, "atualizados": 0, "sem_chave": 0, "nao_encontrados": 0, "erros": 0}
    report = []

    for json_path in json_files:
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception as e:
            stats["erros"] += 1
            report.append(f"  [ERRO JSON] {json_path.name}: {e}")
            continue

        key = json_to_key(data)
        if key is None:
            stats["sem_chave"] += 1
            report.append(f"  [SEM CHAVE] {json_path.name}: campos ausentes/nao mapeáveis "
                          f"(model={data.get('model')}, device={data.get('device')}, "
                          f"size={data.get('model_size')}, lang={data.get('language')}, "
                          f"dur={data.get('audio_duration_s')})")
            continue

        row = key_index.get(key)
        if row is None:
            stats["nao_encontrados"] += 1
            report.append(f"  [nao MAPEADO] {json_path.name}: chave nao encontrada na planilha -> {key}")
            continue

        # Determinar se é inserçao ou atualizaçao
        already_has_data = ws.cell(row, COL_TEMPO).value is not None
        action = "atualizado" if already_has_data else "inserido"

        tempo = safe_round(data.get("processing_time_s"), 4)
        rtf   = extract_rtf(data)
        wer   = safe_round(data.get("wer"), 4)
        ram   = safe_round(data.get("ram_peak_mb"), 2)
        vram  = safe_round(data.get("vram_peak_mb"), 2)

        if not dry_run:
            ws.cell(row, COL_TEMPO).value = tempo
            ws.cell(row, COL_RTF).value   = rtf
            ws.cell(row, COL_WER).value   = wer
            ws.cell(row, COL_RAM).value   = ram
            ws.cell(row, COL_VRAM).value  = vram

        stats["atualizados" if already_has_data else "inseridos"] += 1
        report.append(
            f"  [{action.upper()}] {json_path.name} -> row {row} {key} | "
            f"tempo={tempo}s rtf={rtf} wer={wer} ram={ram}MB vram={vram}MB"
        )

    # Gravar e salvar após processar tudo
    if not dry_run and (stats["inseridos"] + stats["atualizados"]) > 0:
        backup_path = backup_workbook(WORKBOOK_PATH)
        print(f"Backup criado: {backup_path.name}")
        wb.save(WORKBOOK_PATH)
        print("Planilha salva.")

    # Sumário
    print("\n--- Relatório ---")
    for line in report:
        print(line)
    print(f"\n--- Resumo ---")
    print(f"  Inseridos:       {stats['inseridos']}")
    print(f"  Atualizados:     {stats['atualizados']}")
    print(f"  Sem chave:       {stats['sem_chave']}")
    print(f"  Nao mapeados:    {stats['nao_encontrados']}")
    print(f"  Erros de leitura:{stats['erros']}")
    total = sum(stats.values())
    print(f"  Total processado:{total}")
    if dry_run:
        print("\n[DRY-RUN] Nenhuma alteraçao foi gravada.")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Importa JSONs de benchmark para resultados_gerais.xlsx",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
    python importar_resultados.py --folder resultados/faster-whisper/cpu
    python importar_resultados.py --folder resultados/faster-whisper
    python importar_resultados.py --folder resultados/faster-whisper/cpu --dry-run
        """,
    )
    parser.add_argument(
        "--folder",
        type=str,
        required=True,
        help="Pasta com os JSONs a importar (busca recursiva dentro dela)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Simula a importaçao sem gravar nada",
    )

    args = parser.parse_args()
    folder = Path(args.folder).expanduser().resolve()

    if not folder.exists():
        parser.error(f"Pasta nao encontrada: {folder}")

    if not WORKBOOK_PATH.exists():
        parser.error(f"Planilha nao encontrada: {WORKBOOK_PATH}")

    import_folder(folder, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
